import time
import requests
import openpyxl
from bs4 import BeautifulSoup
from urllib.parse import quote

# ================= НАСТРОЙКИ =================
LINKS_PATH = "Links.xlsx"
RESULT_PATH = "result.xlsx"
PAUSE = 0.1
# ============================================

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept-Language': 'en-US,en;q=0.9',
    'Referer': 'https://www.vesselfinder.com/'
}

# Excel
def read_links_from_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    links = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            links.append(str(row[0]).strip())
    return links

def write_results_to_excel(results, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Название', 'IMO', 'MMSI', 'Тип'])

    for row in results:
        ws.append(row)

    wb.save(path)

# Парсинг карточки судна
def extract_imo_mmsi(vsoup):
    imo = ''
    mmsi = ''

    tables = vsoup.select('table.details, table.aparams')

    for table in tables:
        for row in table.find_all('tr'):
            cells = [td.text.strip() for td in row.find_all('td')]
            if len(cells) != 2:
                continue

            label = cells[0]
            val = cells[1].replace(' ', '')

            # формат IMO/MMSI
            if ('IMO' in label or 'MMSI' in label) and '/' in val:
                parts = val.split('/')
                if len(parts) == 2:
                    imo, mmsi = parts
            else:
                if 'IMO' in label and not imo:
                    imo = val
                if 'MMSI' in label and not mmsi:
                    mmsi = val

    return imo, mmsi

# Основная логика
def get_vessel_data(link):
    # кодирование имени (если поиск по name=)
    if 'name=' in link:
        base, name = link.split('name=', 1)
        link = base + 'name=' + quote(name)

    # 1. Запрос страницы поиска
    response = requests.get(link, headers=HEADERS)
    soup = BeautifulSoup(response.text, 'html.parser')

    # 2. Таблица результатов поиска VesselFinder
    rows = soup.select('table.results tbody tr')

    # 0 судов
    if len(rows) == 0:
        return None

    # >1 судна
    if len(rows) > 1:
        return None

    # ровно 1 судно
    row = rows[0]
    ship_link = row.select_one('a.ship-link')

    if not ship_link:
        return None

    vessel_url = 'https://www.vesselfinder.com' + ship_link.get('href')

    # Название и тип из строки поиска
    name = ship_link.select_one('.slna').get_text(strip=True) if ship_link.select_one('.slna') else ''
    vessel_type = ship_link.select_one('.slty').get_text(strip=True) if ship_link.select_one('.slty') else ''

    # 3. Переход на страницу судна
    vresponse = requests.get(vessel_url, headers=HEADERS)
    vsoup = BeautifulSoup(vresponse.text, 'html.parser')

    # 4. Парсинг IMO / MMSI
    imo, mmsi = extract_imo_mmsi(vsoup)

    return [name, imo, mmsi, vessel_type]

def main():
    links = read_links_from_excel(LINKS_PATH)

    results = []
    skipped = 0
    errors = 0

    for i, link in enumerate(links, 1):
        print(f"[{i}/{len(links)}] {link}")

        try:
            data = get_vessel_data(link)

            if data:
                print(f"   Добавлено: {data}")
                results.append(data)
            else:
                print("  ⏭ Пропущено (0 или >1 судна)")
                skipped += 1

        except Exception as e:
            print(f"   Ошибка: {e}")
            errors += 1

        time.sleep(PAUSE)

    write_results_to_excel(results, RESULT_PATH)

    print("\n================= ГОТОВО =================")
    print(f"Добавлено судов : {len(results)}")
    print(f"Пропущено       : {skipped}")
    print(f"Ошибок          : {errors}")
    print(f"Файл результата : {RESULT_PATH}")
    print("==========================================")

if __name__ == "__main__":
    main()
